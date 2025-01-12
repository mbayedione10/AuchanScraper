import pandas as pd
from typing import List, Dict
import io
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

class OdooDataProcessor:
    ODOO_COLUMNS = [
        'name', 'default_code', 'list_price', 'type', 'categ_id',
        'description_sale', 'active', 'sale_ok', 'purchase_ok',
        'image_1920', 'is_published', 'public_categ_ids'
    ]

    COLUMN_WIDTHS = {
        'name': 40,
        'default_code': 15,
        'list_price': 15,
        'type': 10,
        'categ_id': 15,
        'description_sale': 50,
        'active': 8,
        'sale_ok': 8,
        'purchase_ok': 10,
        'image_1920': 50,
        'is_published': 12,
        'public_categ_ids': 30
    }

    def format_data(self, products: List[Dict]) -> pd.DataFrame:
        """Format scraped data into Odoo-compatible DataFrame"""
        # Create DataFrame from products
        df = pd.DataFrame(products)

        # Ensure all required columns exist
        for col in self.ODOO_COLUMNS:
            if col not in df.columns:
                df[col] = ''

        # Convert boolean values to integers (Odoo format)
        bool_columns = ['active', 'sale_ok', 'purchase_ok', 'is_published']
        for col in bool_columns:
            df[col] = 1  # Set all to True by default

        # Format category IDs
        if 'public_categ_ids' in df.columns:
            df['public_categ_ids'] = df['public_categ_ids'].apply(
                lambda x: ','.join(x) if isinstance(x, list) else str(x)
            )

        # Format prices to 2 decimal places
        if 'list_price' in df.columns:
            df['list_price'] = df['list_price'].round(2)

        # Reorder columns to match ODOO_COLUMNS
        return df[self.ODOO_COLUMNS]

    def export_to_excel(self, df: pd.DataFrame) -> bytes:
        """Export DataFrame to Excel bytes with proper formatting"""
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Products')

            # Get the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Products']

            # Define styles
            header_fill = PatternFill(start_color="E40019", end_color="E40019", fill_type="solid")
            header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format headers
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

                # Set column width
                column_letter = cell.column_letter
                worksheet.column_dimensions[column_letter].width = self.COLUMN_WIDTHS.get(column, 15)

            # Format data cells
            for row in range(2, len(df) + 2):
                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=row, column=col_num)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

                    # Special formatting for price column
                    if column == 'list_price':
                        cell.number_format = '#,##0.00 "CFA"'

                    # Center align boolean columns
                    elif column in ['active', 'sale_ok', 'purchase_ok', 'is_published']:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # Freeze the header row
            worksheet.freeze_panes = 'A2'

        return output.getvalue()